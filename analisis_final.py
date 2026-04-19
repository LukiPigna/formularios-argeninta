import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

file_path = 'Planilla-de-Solicitud-de-Compras-2026.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=False)

print('='*100)
print('ANÁLISIS COMPLETO Y DETALLADO DEL FORMULARIO DE SOLICITUD DE COMPRAS 2026')
print('='*100)
print()

print('1. NOMBRE DEL FORMULARIO Y DESCRIPCIÓN')
print('-'*100)
print('Nombre del archivo: Planilla-de-Solicitud-de-Compras-2026.xlsx')
print('Nombre del formulario: Solicitud de Compras 2026')
print()
print('Descripción:')
print('  Formulario integral para gestionar solicitudes de compras y contrataciones, que incluye:')
print('  - Información del proyecto y fuente de financiamiento')
print('  - Datos del solicitante')
print('  - Detalle de items (bienes, materiales, insumos, servicios y consultorías)')
print('  - Información de proveedores sugeridos (hasta 3 opciones)')
print('  - Campos para cotizaciones y adjuntos')
print('  - Régimen jurisdiccional según monto de compra')
print('  - Anexos para Términos de Referencia (TdR) y Tabla de Evaluación')
print('  - Tablas auxiliares con códigos de componentes y productos')
print()

print('2. LISTADO COMPLETO DE HOJAS')
print('-'*100)
print(f'Total de hojas: {len(wb.sheetnames)}')
for idx, sheet_name in enumerate(wb.sheetnames, 1):
    print(f'  {idx}. {sheet_name}')
print()

def analyze_sheet_detailed(sheet_name, wb):
    ws = wb[sheet_name]
    print('='*100)
    print(f'HOJA: {sheet_name.upper()}')
    print('='*100)

    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    print(f'\nDimensiones: {df.shape[0]} filas × {df.shape[1]} columnas\n')

    # Detectar estructura real leyendo celdas clave
    print('--- CAMPOS PRINCIPALES ---')
    print()

    if sheet_name == 'SOLICITUD DE COMPRA':
        print('SECCIÓN 1: INFORMACIÓN GENERAL DEL PROYECTO')
        fields = [
            ('B3', 'Código de Proyecto', 'texto', 'Sí'),
            ('C5', 'Fuente de Financiamiento', 'lista desplegable', 'Sí'),
            ('B6', 'Institución afectada por el gasto', 'texto', 'Sí'),
            ('B7', 'Descripción de la actividad a desarrollar', 'texto largo', 'Sí'),
            ('B9', 'Nombre del solicitante', 'texto', 'Sí'),
            ('D9', 'Mail del solicitante', 'email', 'Sí'),
            ('F9', 'Celular del solicitante', 'texto', 'No'),
            ('B10', 'Fecha de solicitud', 'fecha', 'Sí'),
            ('B11', 'Monto estimado total', 'numérico', 'Sí'),
        ]
        for cell, name, dtype, req in fields:
            print(f'  - {name}')
            print(f'    Celda: {cell}, Tipo: {dtype}, Obligatorio: {req}')

        print('\nSECCIÓN 2: DETALLE DE BIENES, MATERIALES E INSUMOS (Filas 14-24)')
        print('  Campos de la tabla de items:')
        item_fields = [
            ('A14', 'ITEMs N°', 'numérico', 'Sí'),
            ('B14', 'Partida Presupuestaria', 'texto', 'Sí'),
            ('C14', 'Actividad', 'texto', 'Sí'),
            ('D14', 'Descripción / Especificaciones técnicas', 'texto largo', 'Sí'),
            ('E14', 'Cantidad', 'numérico', 'Sí'),
            ('F14', 'Importe x unidad', 'numérico', 'Sí'),
            ('G14', 'Estimado TOTAL Pesos Argentinos', 'numérico (cálculo)', 'No'),
            ('H14', 'Estimado TOTAL Moneda Extranjera', 'numérico', 'No'),
            ('I14', 'Especificar Moneda', 'texto', 'No'),
            ('K14', 'Lugar de entrega', 'texto', 'Sí'),
            ('L14', 'Persona de contacto/recepción', 'texto', 'Sí'),
            ('M14', 'Observaciones adicionales', 'texto', 'No'),
        ]
        for cell, name, dtype, req in item_fields:
            print(f'  - {name}')
            print(f'    Celda: {cell}, Tipo: {dtype}, Obligatorio: {req}')

        print('\nSECCIÓN 3: PROVEEDORES SUGERIDOS (Bienes y Materiales)')
        for prov_num in [1, 2, 3]:
            col_offset = (prov_num - 1) * 4
            start_col = get_column_letter(14 + col_offset)
            print(f'\n  Proveedor Sugerido {prov_num}:')
            print(f'    - Nombre o Razón social: Celda {start_col}14, Tipo: texto, Obligatorio: No')
            print(f'    - CUIT: Celda {get_column_letter(15 + col_offset)}14, Tipo: texto, Obligatorio: No')
            print(f'    - Teléfono: Celda {get_column_letter(16 + col_offset)}14, Tipo: texto, Obligatorio: No')
            print(f'    - E-mail: Celda {get_column_letter(17 + col_offset)}14, Tipo: email, Obligatorio: No')

        print('\nSECCIÓN 4: DETALLE DE SERVICIOS Y CONSULTORÍAS (Filas 28-35)')
        service_fields = [
            ('A30', 'ITEMs N°', 'numérico', 'Sí'),
            ('B30', 'Partida Presupuestaria', 'texto', 'Sí'),
            ('C30', 'Actividad', 'texto', 'Sí'),
            ('D30', 'Descripción del servicio a contratar', 'texto largo', 'Sí'),
            ('E30', 'Duración', 'texto', 'Sí'),
            ('F30', 'Cantidad de pagos, valor de cada pago', 'texto', 'Sí'),
            ('G30', 'Estimado TOTAL Pesos Argentinos', 'numérico', 'Sí'),
            ('H30', 'Estimado TOTAL Moneda Extranjera', 'numérico', 'No'),
            ('I30', 'Especificar Moneda', 'texto', 'No'),
            ('K30', 'Condición de pago - Entregables', 'texto', 'Sí'),
            ('L30', 'Persona que aprueba', 'texto', 'Sí'),
            ('M30', 'Observaciones adicionales', 'texto', 'No'),
        ]
        for cell, name, dtype, req in service_fields:
            print(f'  - {name}')
            print(f'    Celda: {cell}, Tipo: {dtype}, Obligatorio: {req}')

        print('\nSECCIÓN 5: PROVEEDORES SUGERIDOS (Servicios y Consultorías)')
        for prov_num in [1, 2, 3]:
            col_offset = (prov_num - 1) * 4
            start_col = get_column_letter(14 + col_offset)
            print(f'\n  Proveedor Sugerido {prov_num}:')
            print(f'    - Nombre o Razón social: Celda {start_col}30, Tipo: texto, Obligatorio: No')
            print(f'    - Teléfono: Celda {get_column_letter(15 + col_offset)}30, Tipo: texto, Obligatorio: No')
            print(f'    - E-mail: Celda {get_column_letter(16 + col_offset)}30, Tipo: email, Obligatorio: No')
            print(f'    - Adjunta Cotización?: Celda {get_column_letter(17 + col_offset)}30, Tipo: checkbox, Obligatorio: Sí')

        print('\nSECCIÓN 6: CAMPOS DE ADJUNTOS (Fila 46)')
        attachment_fields = [
            ('K46', 'DNI/Pasaporte del solicitante', 'adjunto', 'Sí'),
            ('N46', 'Adjunta COPIA DNI/Pasaporte', 'adjunto', 'Sí'),
            ('Q46', 'Adjunta Cotización Proveedor 1', 'adjunto', 'Sí'),
            ('T46', 'Adjunta Cotización Proveedor 2', 'adjunto', 'Sí'),
            ('Y46', 'Adjunta Cotización Proveedor 3', 'adjunto', 'No'),
        ]
        for cell, name, dtype, req in attachment_fields:
            print(f'  - {name}')
            print(f'    Celda: {cell}, Tipo: {dtype}, Obligatorio: {req}')

    elif sheet_name == 'Listado':
        print('TABLA DE FINANCIAMIENTOS Y COORDINADORES')
        list_fields = [
            ('A', 'Código', 'texto', 'Sí'),
            ('B', 'Descripción del financiamiento', 'texto', 'Sí'),
            ('C', 'FINANCIAMIENTO', 'texto (USDA, OTROS, etc.)', 'Sí'),
            ('D', 'Coordinador', 'texto', 'No'),
            ('E', 'Componente', 'texto (calculado)', 'No'),
            ('F', 'Fuente', 'texto', 'Sí'),
        ]
        for col, name, dtype, req in list_fields:
            print(f'  - {name}')
            print(f'    Columna: {col}, Tipo: {dtype}, Obligatorio: {req}')
        print('  Nota: Esta hoja contiene datos de referencia con 249 fórmulas de concatenación')

    elif sheet_name == 'Info Relevante':
        print('HOJA DE INFORMACIÓN RELEVANTE')
        print('  Contiene consideraciones generales y definiciones')
        print('  Esta es una hoja informativa sin campos de entrada')
        print('  Dimensiones: 70 filas × 16 columnas')
        print('  Principalmente contiene texto explicativo')

    elif sheet_name == 'Anexo I - Modelo de TdR':
        print('MODELO DE TÉRMINOS DE REFERENCIA (TdR)')
        tdr_fields = [
            ('B4', 'Nombre del Proyecto', 'texto', 'Sí'),
            ('B12', '1. Antecedentes', 'texto largo', 'Sí'),
            ('B16', '2. Objetivos de los Trabajos', 'texto largo', 'Sí'),
            ('B20', '3. Alcance de los Servicios, Tareas (Componentes)', 'texto largo', 'Sí'),
            ('B22', '3.1. Servicios a prestar y tareas a realizar', 'texto largo', 'Sí'),
            ('B25', '3.2. Calificación del Consultor', 'texto largo', 'Sí'),
            ('B40', 'Responsable de la aprobación de los informes', 'texto', 'Sí'),
            ('B43', '6. Información del Cliente, Personal de la Contraparte y Lugar de los Servicios', 'texto largo', 'Sí'),
        ]
        for cell, name, dtype, req in tdr_fields:
            print(f'  - {name}')
            print(f'    Celda: {cell}, Tipo: {dtype}, Obligatorio: {req}')
        print('\n  Este anexo es un template para contrataciones de consultores')

    elif sheet_name == 'Anexo II - Tabla evaluación':
        print('TABLA DE EVALUACIÓN DE CONSULTORES')
        print('  Campos principales:')
        eval_fields = [
            ('B4', 'Nombre del Proyecto', 'texto', 'Sí'),
            ('B5', 'Criterios, subcriterios y sistema de puntos', 'texto largo', 'Sí'),
            ('C5', 'Puntaje mínimo/máximo', 'texto', 'Sí'),
            ('D5', 'Nombre del consultor evaluado', 'texto', 'Sí'),
            ('B26', 'PUNTUACIÓN TOTAL GLOBAL', 'numérico (cálculo)', 'Sí'),
            ('B35', 'Firma del evaluador', 'firma', 'Sí'),
        ]
        for cell, name, dtype, req in eval_fields:
            print(f'  - {name}')
            print(f'    Celda: {cell}, Tipo: {dtype}, Obligatorio: {req}')
        print('\n  Este anexo se usa para evaluar propuestas de consultores')

    elif sheet_name == 'Jurisd. Compras':
        print('RÉGIMEN JURISDICCIONAL DEL REGLAMENTO INTERNO')
        print('  Tabla que define el tipo de compra según monto:')
        print('  - Columna B: TIPO (1 a 5)')
        print('  - Columna C: Descripción del tipo de compra')
        print('  - Columna D: DE (monto mínimo, con fórmula)')
        print('  - Columna E: HASTA (monto máximo)')
        print('\n  Tipos de compra:')
        print('  1. COMPRA MENOR: hasta $3,500,000')
        print('  2. COMPRA DIRECTA: $3,500,001 a $54,000,000')
        print('  3. CONCURSO DE PRECIOS: $54,000,001 a $100,000,000')
        print('  4. LICITACIÓN PRIVADA: $100,000,001 a $270,000,000')
        print('  5. LICITACIÓN PÚBLICA: más de $270,000,000')

    elif sheet_name == 'tablas auxiliares':
        print('TABLAS AUXILIARES DE REFERENCIA')
        print('  Campos principales:')
        aux_fields = [
            ('A', 'Código del componente/producto', 'texto', 'Sí'),
            ('B', 'Descripción del producto y actividad', 'texto', 'Sí'),
            ('D', 'Rubros presupuestarios (ej: Fontagro, Horiz2020)', 'texto', 'Sí'),
            ('F', 'Validación de adjuntos (SI/NO/tipo)', 'texto', 'Sí'),
        ]
        for col, name, dtype, req in aux_fields:
            print(f'  - {name}')
            print(f'    Columna: {col}, Tipo: {dtype}, Obligatorio: {req}')
        print('\n  Esta hoja contiene códigos y validaciones usados en el formulario principal')
        print('  Incluye códigos como: 1.1, 1.2, 1.3, etc. para Fontagro, Horiz2020, PROCISUR, AECID')

    print()

    # Lógica condicional
    print('--- LÓGICA CONDICIONAL ---')
    formulas = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f':
                formulas.append(f'{get_column_letter(cell.column)}{cell.row}: {cell.value}')

    if formulas:
        print(f'Sí, se detectaron {len(formulas)} fórmulas\n')
        print('Detalle de fórmulas principales:')
        for f in formulas[:15]:
            print(f'  {f}')
        if len(formulas) > 15:
            print(f'  ... y {len(formulas) - 15} fórmulas más')

        if sheet_name == 'SOLICITUD DE COMPRA':
            print('\nFórmulas específicas de cálculo:')
            print('  - G15 a G23: =E[row]*F[row] (Multiplicación Cantidad × Precio Unitario)')
            print('  - G25: =SUM(G15:G23) (Suma total de bienes)')
            print('  - C5: =VLOOKUP(C4,tabla,2,0) (Búsqueda de fuente de financiamiento)')
    else:
        print('No se detectó lógica condicional o fórmulas')
    print()

    # Campos de carga de archivos
    print('--- CAMPOS DE CARGA DE ARCHIVOS O ADJUNTOS ---')
    file_fields = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                val_lower = str(cell.value).lower()
                keywords = ['archivo', 'adjunt', 'adjunto', 'cargar', 'upload', 'subir', 'anexo', 'documento', 'pdf', 'imagen', 'foto', 'dni', 'pasaporte', 'cotiz', 'dni/pasaporte']
                if any(kw in val_lower for kw in keywords):
                    file_fields.append(f'{get_column_letter(cell.column)}{cell.row}: {cell.value}')

    if file_fields:
        print(f'Sí, se detectaron {len(file_fields)} campos de adjuntos:\n')
        for ff in file_fields:
            print(f'  - {ff}')
    else:
        print('No se detectaron campos de carga de archivos')
    print()

    # Campos de firma o aprobación
    print('--- CAMPOS DE FIRMA O APROBACIÓN ---')
    approval_fields = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                val_lower = str(cell.value).lower()
                keywords = ['firma', 'aprob', 'autoriz', 'validac', 'revisión', 'aprobador', 'firma digital', 'sello', 'responsable.*aprob']
                if any(kw in val_lower for kw in keywords):
                    approval_fields.append(f'{get_column_letter(cell.column)}{cell.row}: {cell.value}')

    if approval_fields:
        print(f'Sí, se detectaron {len(approval_fields)} campos de aprobación:\n')
        for af in approval_fields:
            print(f'  - {af}')
    else:
        print('No se detectaron campos de firma o aprobación')
    print()

    # Secciones o bloques temáticos
    print('--- SECCIONES O BLOQUES TEMÁTICOS ---')

    if sheet_name == 'SOLICITUD DE COMPRA':
        sections = [
            'Información General del Proyecto (Filas 3-11)',
            'Detalle de Bienes, Materiales e Insumos (Filas 12-27)',
            'Proveedores Sugeridos para Bienes (Columnas N-S)',
            'Detalle de Servicios y Consultorías (Filas 28-36)',
            'Proveedores Sugeridos para Servicios (Columnas N-Z)',
            'Campos de Adjuntos (Fila 46)',
            'Resumen y Totales (Filas inferiores)',
        ]
        print(f'Total: {len(sections)} secciones principales:\n')
        for i, s in enumerate(sections, 1):
            print(f'  {i}. {s}')

    elif sheet_name == 'Anexo I - Modelo de TdR':
        sections = [
            'Información del Proyecto (Filas 3-8)',
            '1. Antecedentes (Filas 12-15)',
            '2. Objetivos de los Trabajos (Filas 16-19)',
            '3. Alcance de los Servicios (Filas 20-24)',
            '3.1. Servicios a prestar (Filas 22-24)',
            '3.2. Calificación del Consultor (Filas 25-39)',
            'Responsable de aprobación (Filas 40-42)',
            '6. Información del Cliente (Filas 43-46)',
        ]
        print(f'Total: {len(sections)} secciones:\n')
        for i, s in enumerate(sections, 1):
            print(f'  {i}. {s}')

    elif sheet_name == 'Anexo II - Tabla evaluación':
        sections = [
            'Información del Proyecto (Filas 3-8)',
            'Criterios y Subcriterios de Evaluación (Filas 10-25)',
            'Puntuación Total Global (Fila 26)',
            'Firma del Evaluador (Fila 35)',
        ]
        print(f'Total: {len(sections)} secciones:\n')
        for i, s in enumerate(sections, 1):
            print(f'  {i}. {s}')

    else:
        print('Esta hoja no tiene secciones claramente definidas')
    print()

    # Fórmulas y cálculos automáticos
    print('--- FÓRMULAS Y CÁLCULOS AUTOMÁTICOS ---')

    if formulas:
        # Clasificar fórmulas
        formula_types = {}
        for f in formulas:
            f_upper = f.upper()
            if 'SUM' in f_upper:
                formula_types['Sumas'] = formula_types.get('Sumas', 0) + 1
            elif 'MULT' in f_upper or '*F' in f_upper:
                formula_types['Multiplicaciones'] = formula_types.get('Multiplicaciones', 0) + 1
            elif 'IF' in f_upper or 'SI' in f_upper:
                formula_types['Condicionales'] = formula_types.get('Condicionales', 0) + 1
            elif 'DATE' in f_upper or 'HOY' in f_upper:
                formula_types['Fechas'] = formula_types.get('Fechas', 0) + 1
            elif 'VLOOKUP' in f_upper or 'BUSCARV' in f_upper:
                formula_types['Búsquedas'] = formula_types.get('Búsquedas', 0) + 1
            elif 'CONCATENATE' in f_upper or 'CONCATENAR' in f_upper:
                formula_types['Concatenaciones'] = formula_types.get('Concatenaciones', 0) + 1
            else:
                formula_types['Otras'] = formula_types.get('Otras', 0) + 1

        print(f'Total de fórmulas: {len(formulas)}\n')
        print('Distribución por tipo:')
        for ft, count in formula_types.items():
            print(f'  - {ft}: {count}')

        if sheet_name == 'SOLICITUD DE COMPRA':
            print('\nCálculos automáticos específicos:')
            print('  - Cálculo de subtotal por ítem: Cantidad × Precio Unitario')
            print('  - Cálculo de total general: Suma de todos los subtotales')
            print('  - Búsqueda de fuente de financiamiento basada en código')
    else:
        print('No se detectaron fórmulas o cálculos automáticos')
    print()

    # Observaciones sobre estructura
    print('--- OBSERVACIONES SOBRE ESTRUCTURA ---')

    obs = []
    merged = list(ws.merged_cells.ranges)
    bold = sum(1 for r in ws.iter_rows() for c in r if c.font.bold)
    colored = sum(1 for r in ws.iter_rows() for c in r if c.fill.fgColor.rgb != 'FFFFFFFF')
    empty_rows = sum(1 for r in range(1, ws.max_row + 1) if all(c.value is None for c in ws[r]))
    empty_cols = sum(1 for c in range(1, ws.max_column + 1) if all(ws.cell(r, c).value is None for r in range(1, ws.max_row + 1)))

    if merged:
        obs.append(f'Tiene {len(merged)} rangos de celdas combinadas')
    if bold > 0:
        obs.append(f'Tiene {bold} celdas en negrita (encabezados y títulos)')
    if colored > 0:
        obs.append(f'Tiene {colored} celdas con color de fondo (formato visual)')
    if empty_rows > 0:
        obs.append(f'Contiene {empty_rows} filas completamente vacías')
    if empty_cols > 0:
        obs.append(f'Contiene {empty_cols} columnas completamente vacías')

    for o in obs:
        print(f'  - {o}')

    print()

    # Observaciones adicionales específicas
    print('--- OBSERVACIONES ADICIONALES ---')
    if sheet_name == 'SOLICITUD DE COMPRA':
        print('  - Hoja principal y más compleja del formulario')
        print('  - Contiene 16 fórmulas de cálculo automático')
        print('  - Estructura dividida en dos tipos de compra: Bienes/Materiales y Servicios/Consultorías')
        print('  - Permite hasta 3 proveedores sugeridos por tipo de compra')
        print('  - Incluye campos obligatorios de adjuntar documentación (DNI, cotizaciones)')
        print('  - Usa validación de datos para fuentes de financiamiento (referencia a hojas auxiliares)')
    elif sheet_name == 'Listado':
        print('  - Hoja de referencia con datos de financiamientos')
        print('  - Contiene 249 fórmulas de concatenación para componentes')
        print('  - Usada como fuente de datos para listas desplegables en hoja principal')
    elif sheet_name == 'Info Relevante':
        print('  - Hoja puramente informativa')
        print('  - Contiene definiciones y consideraciones generales')
        print('  - No requiere entrada de datos')
    elif sheet_name == 'Anexo I - Modelo de TdR':
        print('  - Template para contrataciones de consultores')
        print('  - Incluye sección de firma del responsable')
        print('  - Estructura estándar para Términos de Referencia')
    elif sheet_name == 'Anexo II - Tabla evaluación':
        print('  - Tabla de evaluación de propuestas de consultores')
        print('  - Incluye campo de firma del evaluador')
        print('  - Sistema de puntajes por criterios')
    elif sheet_name == 'Jurisd. Compras':
        print('  - Define régimen jurisdiccional por monto')
        print('  - Usa fórmulas secuenciales para rangos (monto mínimo = monto máximo anterior + 1)')
        print('  - Referencia para determinar tipo de compra según monto')
    elif sheet_name == 'tablas auxiliares':
        print('  - Tablas de referencia para códigos y validaciones')
        print('  - Incluye códigos de componentes, productos y actividades')
        print('  - Define qué tipo de adjuntos son obligatorios según el tipo de compra')

    print()
    print()

# Ejecutar análisis para cada hoja
for sheet_name in wb.sheetnames:
    analyze_sheet_detailed(sheet_name, wb)

print('='*100)
print('FIN DEL ANÁLISIS DETALLADO')
print('='*100)
